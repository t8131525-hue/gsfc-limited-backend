import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.shortcuts import get_object_or_404
from django.utils import timezone
from datetime import timedelta
from django.db.models import Avg, Min, Max
from django.db.models.functions import TruncDate
from django.http import HttpResponse

from ..models import Product, TestResult, TestRecord
from ..serializers import ProductQualityDetailSerializer


class ProductQualityDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, product_id, *args, **kwargs):
        product = get_object_or_404(Product, pk=product_id)
        active_version = product.versions.filter(is_active=True).first()

        if not active_version:
            return Response(
                {"error": "This product does not have an active version."},
                status=status.HTTP_404_NOT_FOUND,
            )

        end_date_str = request.query_params.get(
            "end_date", timezone.now().strftime("%Y-%m-%d")
        )
        start_date_str = request.query_params.get(
            "start_date", (timezone.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        )

        trends_data = []
        grades_data = []
        has_grades = active_version.grades.exists()

        if has_grades:
            for grade in active_version.grades.all():
                grade_trends = []
                for param in grade.parameters.all():
                    daily_aggregates = (
                        TestResult.objects.filter(
                            parameter=param,
                            test_record__version=active_version,
                            test_record__product_grade=grade,
                            test_record__created_at__date__range=[
                                start_date_str,
                                end_date_str,
                            ],
                            test_record__status__in=["APPROVED", "CLOSED"],
                        )
                        .annotate(date=TruncDate("test_record__created_at"))
                        .order_by("date")
                        .values("date")
                        .annotate(
                            avg=Avg("value_decimal"),
                            min=Min("value_decimal"),
                            max=Max("value_decimal"),
                        )
                    )
                    param.data_points = list(daily_aggregates)
                    grade_trends.append(param)
                grade.trends = grade_trends
                grades_data.append(grade)
        else:
            for param in active_version.parameters.all():
                daily_aggregates = (
                    TestResult.objects.filter(
                        parameter=param,
                        test_record__version=active_version,
                        test_record__created_at__date__range=[
                            start_date_str,
                            end_date_str,
                        ],
                        test_record__status__in=["APPROVED", "CLOSED"],
                    )
                    .annotate(date=TruncDate("test_record__created_at"))
                    .order_by("date")
                    .values("date")
                    .annotate(
                        avg=Avg("value_decimal"),
                        min=Min("value_decimal"),
                        max=Max("value_decimal"),
                    )
                )
                param.data_points = list(daily_aggregates)
                trends_data.append(param)

        recent_tests = TestRecord.objects.filter(version__product=product).order_by(
            "-created_at"
        )[:10]

        if request.query_params.get("format") == "excel":
            # ✅ 1. Pass the start and end dates to the export function
            return self.export_to_excel(
                product,
                active_version,
                trends_data,
                grades_data,
                recent_tests,
                start_date_str,
                end_date_str,
            )

        context_data = {
            "id": product.id,
            "name": product.name,
            "product_id": product.product_id,
            "active_version": active_version,
            "has_grades": has_grades,
            "grades": grades_data,
            "trends": trends_data,
            "recent_tests": recent_tests,
        }

        serializer = ProductQualityDetailSerializer(context_data)
        return Response(serializer.data)

    # ✅ 2. Update the method signature to accept the dates
    def export_to_excel(
        self,
        product,
        active_version,
        trends_data,
        grades_data,
        recent_tests,
        start_date,
        end_date,
    ):
        workbook = openpyxl.Workbook()

        # --- Summary Sheet ---
        summary_ws = workbook.active
        summary_ws.title = "Summary"
        summary_ws.append(["Product Name", product.name])
        summary_ws.append(["Product ID", product.product_id])
        summary_ws.append(["Active Version", active_version.version_name])
        summary_ws.append(["Date Range", f"{start_date} to {end_date}"])
        summary_ws.append(
            ["Report Generated On", timezone.now().strftime("%Y-%m-%d %H:%M:%S")]
        )

        # --- Trends Data Sheet(s) ---
        if grades_data:
            for grade in grades_data:
                ws = workbook.create_sheet(title=f"Trends - {grade.name[:20]}")
                self._write_trends_to_sheet(ws, grade.trends)
        else:
            ws = workbook.create_sheet(title="Trends Data")
            self._write_trends_to_sheet(ws, trends_data)

        # --- Recent Tests Sheet ---
        recent_tests_ws = workbook.create_sheet(title="Recent Tests")
        recent_tests_ws.append(["Record ID", "Status", "Analyst", "Date"])
        for test in recent_tests:
            analyst_name = (
                test.test_record.analyst.get_full_name()
                if test.test_record.analyst
                else "N/A"
            )
            naive_datetime = timezone.make_naive(test.test_record.created_at)
            recent_tests_ws.append(
                [
                    test.test_record.record_id,
                    test.test_record.status,
                    analyst_name,
                    naive_datetime,
                ]
            )

        # Prepare and return the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        safe_product_name = "".join(
            c for c in product.name if c.isalnum() or c in (" ", "_")
        ).rstrip()

        # ✅ 3. Construct the new dynamic filename
        filename = f"{safe_product_name}_{start_date}_to_{end_date}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response

    def _write_trends_to_sheet(self, worksheet, trends):
        for param in trends:
            worksheet.append(
                [
                    f"Parameter: {param.name}",
                    f"Spec Range: {param.min_value or '-'} to {param.max_value or '-'}",
                ]
            )
            worksheet.append(["Date", "Average", "Minimum", "Maximum"])
            for dp in param.data_points:
                worksheet.append([dp["date"], dp["avg"], dp["min"], dp["max"]])
            worksheet.append([])
