import base64
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework import viewsets, permissions
from ..models import TestRecord
from django_filters.rest_framework import DjangoFilterBackend
from ..serializers import TestRecordSerializer
from product_testing_system.pagination import StandardResultsSetPagination
from rest_framework.filters import SearchFilter, OrderingFilter
from ..filters import TestRecordFilter
from django.utils import timezone
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework import viewsets, permissions, status
from ..serializers import (
    TestRecordSerializer,
    RecentTestRecordSerializer,
)
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.utils import timezone
from django.db import transaction
from django.contrib.auth import get_user_model
from audit_trail.utils import log_custom_event
from ..serializers.AssignAnalystSerializer import AssignAnalystSerializer
from weasyprint import HTML
from django.conf import settings


User = get_user_model()


class TestRecordViewSet(viewsets.ModelViewSet):
    queryset = TestRecord.objects.all()
    serializer_class = TestRecordSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = TestRecordFilter
    search_fields = [
        "sample_id",
        "batch_no",
        "version__product__name",
        "record_id",
        "analyst__username",
        "analyst__first_name",
        "analyst__last_name",
    ]
    ordering_fields = ["created_at", "analyst__username", "status", "lab__name"]
    ordering = ["-created_at"]

    def get_queryset(self):
        """
        Dynamically filters the queryset.
        - Non-managers only see their own records.
        - The 'Recent' view is automatically filtered for today's date.
        """
        user = self.request.user
        view_type = self.request.query_params.get("view_type", "recent")

        queryset = TestRecord.objects.select_related(
            "version__product", "product_grade", "analyst", "lab"
        ).all()

        if not user.has_perm("inventory.can_view_all_test_records"):
            queryset = queryset.filter(analyst=user)

        if self.action == "list" and view_type == "recent":
            today = timezone.now().date()
            queryset = queryset.filter(created_at__date=today)

        return queryset

    def get_serializer_class(self):
        """
        Chooses the serializer based on the action and view type.
        - 'list' action gets a lightweight serializer.
        - Other actions (retrieve, create, update) get the full serializer.
        """
        if self.action == "list":
            view_type = self.request.query_params.get("view_type", "recent")
            if view_type == "historical":
                return RecentTestRecordSerializer
            return RecentTestRecordSerializer

        return TestRecordSerializer

    @action(
        detail=True,
        methods=["get"],
        url_path="download-pdf",
        permission_classes=[permissions.IsAuthenticated],
    )
    def download_pdf(self, request, pk=None):
        """
        Generates and returns a PDF report for a specific test record.
        """
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance, context={"request": request})
            record_data = serializer.data
            logo_base64 = None
            logo_path = os.path.join(settings.BASE_DIR, "product_testing_system", "static", "images", "logo.png")

            try:
                with open(logo_path, "rb") as image_file:
                    logo_base64 = base64.b64encode(image_file.read()).decode("utf-8")
            except FileNotFoundError:
                print(f"Logo file not found at {logo_path}")
                pass 

            context = {
                "record": instance,
                "data": record_data,  # 👈 Pass the serialized data for the table/signatures
                "logo_base64": logo_base64,
            }

            html_string = render_to_string("reports/test_record_report.html", context)
            pdf_file = HTML(
                string=html_string, base_url=request.build_absolute_uri()
            ).write_pdf()
            response = HttpResponse(pdf_file, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="Report-{instance.record_id}.pdf"'
            )

            return response

        except Exception as e:
            # Add this for better debugging in your console
            import traceback

            traceback.print_exc()

            return Response(
                {"error": f"Failed to generate PDF: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    # 👇 =================================================================
    # 👇   ADD THIS NEW FUNCTION
    # 👇 =================================================================
    @action(
        detail=True,
        methods=["get"],
        url_path="download-excel",
        permission_classes=[permissions.IsAuthenticated],
    )
    def download_excel(self, request, pk=None):
        """
        Generates and returns an Excel (.xlsx) report for a specific test record.
        """
        
        def get_expected_range(param_data):
            data_type = param_data.get("data_type")
            
            if data_type in ["INTEGER", "DECIMAL"]:
                min_val = param_data.get("min_value")
                max_val = param_data.get("max_value")
                if min_val is not None and max_val is not None:
                    return f"{min_val} - {max_val}"
                if min_val is not None:
                    return f">= {min_val}"
                if max_val is not None:
                    return f"<= {max_val}"
                return "NA" # Changed from N/A
                
            elif data_type == "ENUM":
                options = param_data.get("enum_options")
                if options and isinstance(options, list):
                    return ", ".join(options)
                return "NA" # Changed from N/A
                
            elif data_type == "BOOLEAN":
                true_label = param_data.get("boolean_true_label", "True")
                false_label = param_data.get("boolean_false_label", "False")
                return f"{true_label} / {false_label}"
                
            elif data_type == "STRING":
                return "Text"
                
            return "NA" # Changed from N/A

        try:
            # 1. Get object and serializer data
            instance = self.get_object() # 👈 We need the raw instance
            serializer = TestRecordSerializer(instance, context={"request": request})
            record_data = serializer.data

            # 2. Create an in-memory workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Report"
            
            bold_font = Font(bold=True)

            # 3. Add Main Record Details
            ws.append(["Test Record Details"])
            ws["A1"].font = bold_font
            ws.merge_cells("A1:B1")

            # --- UPDATED SECTION ---
            # We now use .get(key) or "NA" to handle blank/null values
            main_details = [
                ("Record ID", record_data.get("record_id") or "NA"),
                ("Product ID", instance.version.product.product_id or "NA"),
                ("Product", record_data.get("product_name") or "NA"),
                ("Testing Version", record_data.get("version_name") or "NA"),
                ("Grade", record_data.get("product_grade_name") or "NA"),
                ("Batch No", record_data.get("batch_no") or "NA"),
                ("Sample ID", record_data.get("sample_id") or "NA"),
                ("Lab", record_data.get("lab_name") or "NA"),
                ("Status", record_data.get("status") or "NA"),
                ("Decision", record_data.get("decision") or "NA"),
                ("Tested At", record_data.get("created_at") or "NA"),
                ("Analyst", record_data.get("analyst_full_name") or "NA"),
                ("Approved By", record_data.get("approved_by_full_name") or "NA"),
                ("Approved At", record_data.get("approved_at") or "NA"),
            ]
            # --- END UPDATED SECTION ---
            
            for row, (key, value) in enumerate(main_details, start=2):
                ws[f"A{row}"] = key
                ws[f"A{row}"].font = bold_font
                ws[f"B{row}"] = value

            # Add a spacer row
            spacer_row = len(main_details) + 3
            ws.append([]) # Add an empty row

            # 4. Add Parameter Results
            results_header_row = spacer_row
            ws[f"A{results_header_row}"] = "Test Results"
            ws[f"A{results_header_row}"].font = bold_font
            ws.merge_cells(f"A{results_header_row}:D{results_header_row}")

            param_headers = ["Parameter", "Expected Range", "Result", "Status"]
            ws.append(param_headers)
            for cell in ws[ws.max_row]: 
                cell.font = bold_font

            param_values = record_data.get("parameter_values", [])
            if param_values:
                for param in param_values:
                    param_data = param.get("parameter", {})
                    expected_range = get_expected_range(param_data)
                    
                    ws.append([
                        param_data.get("name") or "NA",
                        expected_range,
                        param.get("display_value") or "NA",
                        param.get("status") or "NA",
                    ])
            else:
                ws.append(["No test results recorded."])

            # 5. Set column widths
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 15

            # 6. Save workbook to an in-memory stream
            with io.BytesIO() as b:
                wb.save(b)
                b.seek(0)

                # 7. Create the HTTP response
                response = HttpResponse(
                    b.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = (
                    f'attachment; filename="Report-{instance.record_id}.xlsx"'
                )
                return response

        except Exception as e:
            return Response(
                {"error": f"Failed to generate Excel: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    # 👇 =================================================================
    # 👇   END OF NEW FUNCTION
    # 👇 =================================================================
    @action(
        detail=True, methods=["patch"], permission_classes=[permissions.IsAuthenticated]
    )
    def assign(self, request, pk=None):
        """
        Assigns an analyst to an unassigned test record.
        Only accessible by users with approval permissions (Managers/Supervisors).
        """
        user = request.user
        if not user.has_perm("inventory.can_approve_test_records"):
            return Response(
                {"detail": "You do not have permission to assign tests."},
                status=status.HTTP_403_FORBIDDEN,
            )

        test_record = self.get_object()

        if test_record.analyst is not None:
            return Response(
                {"detail": "This test has already been assigned."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = AssignAnalystSerializer(data=request.data)
        if serializer.is_valid():
            analyst_id = serializer.validated_data["analyst_id"]
            analyst_to_assign = User.objects.get(pk=analyst_id)

            test_record.analyst = analyst_to_assign
            test_record.save()

            log_custom_event(
                instance=test_record,
                action_type="ASSIGNED",
                user=user,
                details=f"Assigned to analyst {analyst_to_assign.username} by {user.username}.",
            )

            return Response(
                self.get_serializer(test_record).data, status=status.HTTP_200_OK
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated]
    )
    def order_retest(self, request, pk=None):
        user = request.user
        if not user.has_perm("inventory.can_approve_test_records"):
            return Response(
                {"detail": "You do not have permission to order a retest."},
                status=status.HTTP_403_FORBIDDEN,
            )

        with transaction.atomic():
            original_test = self.get_object()
            if original_test.status not in ["APPROVED", "REJECTED"]:
                return Response(
                    {
                        "detail": "Can only order a retest for a record that is 'APPROVED' or 'REJECTED'."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            serializer = AssignAnalystSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            analyst_id = serializer.validated_data["analyst_id"]
            try:
                analyst_to_assign = User.objects.get(pk=analyst_id)
            except User.DoesNotExist:
                return Response(
                    {"detail": "Selected analyst not found."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            original_test.status = "RETEST_ORDERED"
            original_test.retest_ordered_by = user
            original_test.retest_ordered_at = timezone.now()
            original_test.save(
                update_fields=["status", "retest_ordered_by", "retest_ordered_at"]
            )

            new_test = TestRecord.objects.create(
                version=original_test.version,
                lab=original_test.lab,
                product_grade=original_test.product_grade,
                batch_no=original_test.batch_no,
                sample_id=original_test.sample_id,
                status="PENDING",
                analyst=analyst_to_assign,
                retest_of=original_test,
            )
            log_custom_event(
                instance=original_test,
                action_type="RETEST_ORDERED",
                user=user,
                details=f"Retest ordered by {user.username} and assigned to {analyst_to_assign.username}. New record: {new_test.record_id}",
            )
        response_serializer = self.get_serializer(new_test)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)

    @action(
        detail=True, methods=["patch"], permission_classes=[permissions.IsAuthenticated]
    )
    def approve_reject(self, request, pk=None):
        """
        Approves or rejects a test record. This is a final action.
        """
        user = request.user
        if not user.has_perm("inventory.can_approve_test_records"):
            return Response(
                {"detail": "You do not have permission to perform this action."},
                status=status.HTTP_403_FORBIDDEN,
            )

        with transaction.atomic():
            test_record = TestRecord.objects.select_for_update().get(pk=pk)

            if test_record.status != "PENDING":
                return Response(
                    {
                        "detail": "This record has already been actioned by another user."
                    },
                    status=status.HTTP_409_CONFLICT,
                )

            new_status = request.data.get("status")
            comments = request.data.get(
                "supervisor_comments", test_record.supervisor_comments
            )

            if new_status not in ["APPROVED", "REJECTED"]:
                return Response(
                    {"error": f"Invalid status '{new_status}'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            test_record.status = new_status
            test_record.supervisor_comments = comments
            test_record.approved_by = user
            test_record.approved_at = timezone.now()
            test_record.decision = new_status  # Set the permanent decision
            test_record.save(
                update_fields=[
                    "status",
                    "supervisor_comments",
                    "approved_by",
                    "approved_at",
                    "decision",
                ]
            )
            log_custom_event(
                instance=test_record,
                action_type=new_status,
                details=f"Record status changed to {new_status} by {user.username}.",
                user=user,
            )

        serializer = self.get_serializer(test_record)
        return Response(serializer.data)

    @action(
        detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated]
    )
    def close_record(self, request, pk=None):
        """
        Closes an APPROVED or REJECTED test record. This is a final state.
        """
        user = request.user
        if not user.has_perm("inventory.can_approve_test_records"):
            return Response(
                {"detail": "You do not have permission to close records."},
                status=status.HTTP_403_FORBIDDEN,
            )

        with transaction.atomic():
            test_record = TestRecord.objects.select_for_update().get(pk=pk)

            if test_record.status not in ["APPROVED", "REJECTED", "RETEST_ORDERED"]:
                return Response(
                    {
                        "detail": f"Cannot close a record with status '{test_record.get_status_display()}'."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            old_status = test_record.status
            test_record.status = "CLOSED"
            test_record.closed_by = request.user
            test_record.closed_at = timezone.now()
            test_record.save(update_fields=["status", "closed_by", "closed_at"])

            log_custom_event(
                instance=test_record,
                action_type="CLOSED",
                user=user,
                details=f"Record status changed from {old_status} to CLOSED by {user.username}.",
            )

        serializer = self.get_serializer(test_record)
        return Response(serializer.data, status=status.HTTP_200_OK)
